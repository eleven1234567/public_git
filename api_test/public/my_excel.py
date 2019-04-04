from openpyxl import load_workbook
from config.read_conf import ReadConf
from public import path_conf

class MyExcel:
    def __init__(self,file,sheet,init_data):
        self.sheet=sheet
        self.init_data=init_data
        self.file=file

    def my_excel(self):
        wb=load_workbook(self.file)
        sheet=wb[self.sheet]
        test_data=[]
        tel=self.get_init_data()
        for i in range(2,sheet.max_row+1):
            dict_1 = {}
            if sheet.cell(i,3).value=='N':
                continue
            dict_1['case_id']=sheet.cell(i,1).value
            dict_1['title']=sheet.cell(i,2).value
            dict_1['method'] = sheet.cell(i,4).value
            dict_1['url'] = ReadConf().read_conf(path_conf.test_conf_path,'IP','ip')+sheet.cell(i,5).value
            if self.sheet=='register' :
                if sheet.cell(i,6).value.find('${no_reg_tel}')!=-1:
                    dict_1['param'] = sheet.cell(i,6).value.replace('${no_reg_tel}',str(tel))
                    no_reg_tel = tel + 1
                    self.updata_data(no_reg_tel)
                else:
                    dict_1['param'] = sheet.cell(i, 6).value
            else:
                dict_1['param'] = sheet.cell(i, 6).value
            dict_1['ExpectedResult'] = sheet.cell(i,7).value
            dict_1['CheckSql'] = sheet.cell(i,8).value
            dict_1['ActualResult'] = sheet.cell(i,9).value
            dict_1['TestResult'] = sheet.cell(i,10).value
            dict_1['SQLResult'] = sheet.cell(i,11).value
            test_data.append(dict_1)
        return test_data

    #将值写入excel中
    def write_data(self,n,r,res):
        wb = load_workbook(self.file)
        sheet = wb[self.sheet]
        sheet.cell(n,r).value=res
        wb.save(self.file)
        wb.close()

    #获取初始值
    def get_init_data(self):
        wb = load_workbook(self.file)
        sheet = wb[self.init_data]
        no_reg_tel=sheet.cell(1,2).value
        wb.close()
        return no_reg_tel

    # #更新初始值
    def updata_data(self,tel):
        wb = load_workbook(self.file)
        sheet = wb[self.init_data]
        sheet.cell(1,2).value=tel
        wb.save(self.file)
        wb.close()





if __name__ == '__main__':
    # print(MyExcel(path_conf.test_case_path,'recharge','init_data').my_excel())
    # print(MyExcel(path_conf.test_case_path,'register','init_data').my_excel())
    print(MyExcel(path_conf.test_case_path, 'add_loan', 'init_data').my_excel())


