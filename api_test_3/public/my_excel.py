from openpyxl import load_workbook
from  public.read_conf import ReadConf
from public import read_path

# 处理数据的整体思路：先读初始化数据 -- 处理数据 -- 对相关数据进行替换 -更新初始化数据

class MyExcel:
    def __init__(self,file_path,sheet_name):
        self.file_path=file_path
        self.sheet_name=sheet_name

    def my_excel(self):
        wb=load_workbook(self.file_path)
        sheet=wb[self.sheet_name]
        test_data=[] #把所有行的数据存到这个列表里面
        init_data=self.get_init_data() #获取初始化用户名
        # new_init_data=init_data.replace(init_data[-1],str(int(init_data[-1])+1))#现有基础上+1  自增
        for i in  range(2,sheet.max_row+1):
            if sheet.cell(i,4).value =='N':
                continue
            sub_data={}#最重要  每一行的数据单独存在一个字典里面
            sub_data['case_id']=sheet.cell(i,1).value
            sub_data['title'] = sheet.cell(i,2).value
            sub_data['method'] = sheet.cell(i,3).value
            sub_data['url'] = ReadConf().read_conf(read_path.config_path,'IP','ip')+sheet.cell(i,5).value
            sub_data['expect'] = sheet.cell(i,8).value #期望值
            sub_data['sql'] = sheet.cell(i,7).value
            if sheet.cell(i,6).value.find('${no_reg_name}') != -1:
                sub_data['param'] = sheet.cell(i,6).value.replace('${no_reg_name}', init_data)
            else:
                sub_data['param'] = sheet.cell(i,6).value
            test_data.append(sub_data)
            # self.updata_data(new_init_data)
        return test_data

    def write_data(self,r,n,actual_result):
        wb=load_workbook(self.file_path)
        sheet=wb[self.sheet_name]
        sheet.cell(r,n).value=actual_result
        wb.save(self.file_path)

    def get_init_data(self):
        wb=load_workbook(self.file_path)
        sheet=wb['init_data']
        value=sheet.cell(1,2).value
        return value

    def updata_data(self,res):
        wb = load_workbook(self.file_path)
        sheet = wb['init_data']
        sheet.cell(1,2).value=res
        wb.save(self.file_path)



if __name__ == '__main__':
    test_data=MyExcel(read_path.testCase_path,'test_gold_add').my_excel()
    print(test_data)
    # init_data=MyExcel('testCase.xlsx','test_data').get_init_data()
    # print(init_data)